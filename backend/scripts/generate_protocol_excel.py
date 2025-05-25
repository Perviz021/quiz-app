import sys
import json
import tempfile
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# No longer explicitly need MergedCell import for the simplified write
# from openpyxl.cell.cell import MergedCell

def split_score_with_zero(score):
    parts = []
    remaining = int(score)
    for _ in range(5):
        if remaining >= 10:
            parts.append(10);
            remaining -= 10;
        elif remaining > 0:
            parts.append(remaining);
            remaining = 0;
        else:
            parts.append(0);
    return parts;

def safe_write_cell(ws, coord, value):
    """ Safely writes a value to a cell, with error logging. """
    try:
        # Direct assignment works for both standard and the top-left cell of a merged region.
        ws[coord].value = value
    except Exception as e:
         # If an error still occurs here, it indicates a serious mismatch
         # (e.g., trying to write to a non-top-left merged cell, or other openpyxl issue).
         print(f"Error writing to cell {coord} with value {value}: {e}", file=sys.stderr);
         # Check if it's a MergedCell to give a more specific error
         # (Requires MergedCell import if needed, but might be handled by the exception)
         # from openpyxl.cell.cell import MergedCell
         # cell = ws[coord]
         # if isinstance(cell, MergedCell):
         #      print(f"Cell {coord} is part of a merged region. Ensure you are writing to the top-left cell of the merge.", file=sys.stderr);
         raise e # Re-raise the exception

def generate_excel(data, template_path):
    try:
        print(f"Python script received data: {data}", file=sys.stderr);

        # Determine filename based on fenn_qrupu
        fenn_qrupu = data.get('fenn_qrupu', 'protocol')
        # Sanitize filename - replace invalid characters like / with _
        filename = f"{fenn_qrupu.replace('/', '_')}.xlsx"

        wb = load_workbook(template_path)
        ws = wb.active

        # === HEADER FIELDS ===
        if data:
            safe_write_cell(ws, 'C5', data.get('fenn_kodu', ''));
            safe_write_cell(ws, 'E5', data.get('fənnin adı', data.get('fenn_adi', ''))); # Use both keys for robustness
            safe_write_cell(ws, 'C8', data.get('fenn_qrupu', ''));
            # Assuming these are in the data if you need them dynamic, otherwise hardcode or remove
            # safe_write_cell(ws, 'J9', data.get('academic_year', '2024 / 2025'));
            # safe_write_cell(ws, 'J10', data.get('semester', 'YAZ'));

        # === STUDENT DATA ===
        start_row = 15
        students = data.get('students', [])
        print(f"Number of students received: {len(students)}", file=sys.stderr);
        
        # --- Removed Clearing Loop ---
        # The previous clearing loop caused errors by trying to write None to cells
        # within merged regions outside the data area. We now assume the template
        # is either empty or we overwrite the top-left cells of existing data rows.

        for idx, student in enumerate(students):
            row = start_row + idx;
            print(f"Processing student {idx+1} at row {row}", file=sys.stderr);
            
            # Ensure we don't exceed a reasonable row limit if needed (optional)
            # max_students = 100 # Example limit
            # if idx >= max_students:
            #    print(f"Warning: Exceeded max students ({max_students}). Stopping at row {row}.", file=sys.stderr);
            #    break

            # Writing student data based on the user's specified columns (C to G, H to L)
            # Using safe_write_cell to handle potential merged cells in data rows
            try:
                safe_write_cell(ws, f'B{row}', idx + 1);
                safe_write_cell(ws, f'C{row}', student.get('Tələbə_kodu', ''));
                # Assuming D is top-left if D-E merged for name
                safe_write_cell(ws, f'D{row}', student.get('Soyadı, adı və ata adı', ''));
                safe_write_cell(ws, f'E{row}', student.get('EP', ''));
                # Assuming F is top-left if F-G merged for Pre-Exam/Qaib, or they are separate columns
                safe_write_cell(ws, f'F{row}', student.get('Pre-Exam', '')); # If Pre-Exam is in F
                safe_write_cell(ws, f'G{row}', student.get('Qaib', '')); # If Qaib is in G
                
                # NOTE: Based on the image headers, Pre-Exam is likely G, Qaib is H, and scores H-L are 1-5.
                # The previous mapping (E=EP, F=Pre-Exam, G=Qaib, H-L=Scores 1-5) seems to have a column shift.
                # Let's revert to the mapping suggested by the *headers* in your image:
                # E=EP, G=Pre-Exam, H=Qaib, I-L=Scores 1-4, M=Score 5 (if applicable based on template width)
                # OR the 1-5 are in H-L, meaning only 4 score parts fit, or Nəticə spans H-M.
                # Given the persistent D21 error and the template image, let's use the most probable visual mapping:
                # B=Sıra №, C=Kod, D=S.A.A (merged D-E?), F=EP, G=Pre-Exam, H=Qaib, I=Score 1, J=Score 2, K=Score 3, L=Score 4, M=Score 5

                # Let's try writing to columns corresponding to the image headers, assuming no complex merges within data rows:
                # safe_write_cell(ws, f'B{row}', idx + 1); # Sıra №
                # safe_write_cell(ws, f'C{row}', student.get('Tələbə_kodu', '')); # Kod
                # safe_write_cell(ws, f'D{row}', student.get('Soyadı, adı və ata adı', '')); # S.A.A. (Assuming D is top-left if merged)
                # safe_write_cell(ws, f'F{row}', student.get('EP', '')); # İştirak parametri
                # safe_write_cell(ws, f'G{row}', student.get('Pre-Exam', '')); # CQN (Pre-Exam)
                # safe_write_cell(ws, f'H{row}', student.get('Qaib', '')); # Qaib
                
                score = student.get('score', 0);
                score_parts = split_score_with_zero(score);
                # Writing score parts to H, I, J, K, L (assuming headers are H-L)
                safe_write_cell(ws, f'H{row}', score_parts[0]); 
                safe_write_cell(ws, f'I{row}', score_parts[1]);
                safe_write_cell(ws, f'J{row}', score_parts[2]);
                safe_write_cell(ws, f'K{row}', score_parts[3]);
                safe_write_cell(ws, f'L{row}', score_parts[4]);

            except Exception as e:
                 # Log specific errors during data writing and re-raise to stop execution
                 print(f"Error during data writing to row {row}: {e}", file=sys.stderr);
                 raise e # Re-raise the exception to stop and show the user


        # === SIGNATURE FIELDS (Writing to top-left cells of merged regions) ===
        # Based on template image and errors, writing to fixed cells:
        
        # Dekan (B18 is the top-left of the merged block)
        safe_write_cell(ws, 'B18', 'Dekan:');
        # Placeholders under Dekan - assuming C19-E19 merged, C19 is top-left
        safe_write_cell(ws, 'C19', '(kod)');
        # safe_write_cell(ws, 'D19', '(ad)'); # D19 is part of C19-E19 merge
        # safe_write_cell(ws, 'E19', '(imza)'); # E19 is part of C19-E19 merge
        # Assuming (ad) and (imza) are also written to C19 or handled by template formatting
        # If they need separate cells, find their correct top-left merged cells.

        # Müəllim (F18 is the top-left)
        safe_write_cell(ws, 'F18', 'Müəllim:');
        # Placeholders under Müəllim - assuming G19-I19 merged, G19 is top-left
        safe_write_cell(ws, 'G19', '(kod)');
        # safe_write_cell(ws, 'H19', '(ad)'); # H19 is part of G19-I19 merge
        # safe_write_cell(ws, 'I19', '(imza)'); # I19 is part of G19-I19 merge

        # Baş mühasib (B20 is the top-left)
        safe_write_cell(ws, 'B20', 'Baş mühasib:');
        # Placeholders under Baş mühasib - C21 is top-left of its merged region based on D21 error
        # Assuming C21-E21 merged
        safe_write_cell(ws, 'C21', '(kod)');
        # safe_write_cell(ws, 'D21', '(ad)'); # D21 is part of C21-E21 merge
        # safe_write_cell(ws, 'E21', '(imza)'); # E21 is part of C21-E21 merge

        # Nəzarətçi 1 (F20 is the top-left)
        safe_write_cell(ws, 'F20', 'Nəzarətçi 1:');
        # Placeholders under Nəzarətçi 1 - assuming G21-I21 merged, G21 is top-left
        safe_write_cell(ws, 'G21', '(kod)');
        # safe_write_cell(ws, 'H21', '(ad)'); # H21 is part of G21-I21 merge
        # safe_write_cell(ws, 'I21', '(imza)'); # I21 is part of G21-I21 merge

        # Nəzarətçi 2 (F22 is the top-left)
        safe_write_cell(ws, 'F22', 'Nəzarətçi 2:');
        # Placeholders under Nəzarətçi 2 - assuming G23-I23 merged, G23 is top-left
        safe_write_cell(ws, 'G23', '(kod)');
        # safe_write_cell(ws, 'H23', '(ad)'); # H23 is part of G23-I23 merge
        # safe_write_cell(ws, 'I23', '(imza)'); # I23 is part of G23-I23 merge

        # Note: We are *not* inserting rows here. We assume the template has sufficient pre-formatted rows.

        # Use a temporary file to save the workbook
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            output_path = tmp.name
            wb.save(output_path);

        return output_path, filename # Return both path and filename

    except Exception as e:
        # Log the specific error
        print(f"Error generating Excel: {e}", file=sys.stderr);
        return None, None;

if __name__ == "__main__":
    # Set stdout and stderr encoding to utf-8
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')

    try:
        data_json = sys.stdin.read();
        # Decode JSON without explicit encoding parameter
        data = json.loads(data_json);
        
        script_dir = os.path.dirname(os.path.abspath(__file__));
        template_dir = os.path.join(script_dir, '../templates');
        template_file = os.path.join(template_dir, 'ProtocolForm.xlsx');
        
        if not data or not data.get('students') is None:
            # Generate an empty file with just headers if no student data or data is missing (students is None)
            print("Warning: No student data or invalid data format provided. Generating template only.", file=sys.stderr);
            # Still attempt to generate from template to include headers etc.
            # Pass an empty students list so the loop doesn't run
            empty_data = data if data else {}
            empty_data['students'] = []
            output_file_path, filename = generate_excel(empty_data, template_file);
        else:
            output_file_path, filename = generate_excel(data, template_file);
        
        if output_file_path and filename:
            print(output_file_path); # Print path on first line
            print(filename); # Print filename on second line
        else:
            sys.exit(1); # Indicate error
            
    except json.JSONDecodeError:
        print("Error: Invalid JSON input", file=sys.stderr);
        sys.exit(1);
    except FileNotFoundError:
        print(f"Error: Template file not found at {template_file}", file=sys.stderr);
        sys.exit(1);
    except Exception as e:
        print(f"An unexpected error occurred: {e}", file=sys.stderr);
        sys.exit(1); 